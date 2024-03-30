from openpyxl import load_workbook

# 打开 Excel 文件
workbook = load_workbook('fin_excel/国际化（更新历史数据）.xlsx')

# 遍历每一行，生成 SQL 插入语句的值部分
values = []

# 遍历每个工作表
for sheet_name in workbook.sheetnames:

    if sheet_name.startswith("菜单") or sheet_name.startswith("导出"):
        print(sheet_name)
        sheet = workbook[sheet_name]

        # 生成 SQL 插入语句的初始部分
        insert_statement = "INSERT INTO `i18n` (`wkey`, `zh`, `en`, `sp`, `ru`, `fr`) VALUES "


        for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始遍历，第一行通常为标题
            # 检查行是否为空
            if not all(cell is None for cell in row):
                key, zh, en, sp, ru, fr, ex = row
                # 处理可能为空的字段
                zh = zh if zh is not None else 'None'
                en = en if en is not None else 'None'
                ru = ru if ru is not None else 'None'
                sp = sp if sp is not None else 'None'
                fr = fr if fr is not None else 'None'

                key = key.strip()

                # 处理单引号，以防在 SQL 语句中引起问题
                fr = fr.replace("'", "''")
                en = en.replace("'", "''")
                ru = ru.replace("'", "''")
                sp = sp.replace("'", "''")
                values.append(insert_statement+f"('{key}', '{zh}', '{en}', '{sp}', '{ru}', '{fr}');")

        # # 将所有值连接成一个字符串
        # sql_statement = "\n".join(values)

        # 打印生成的 SQL 语句
        # print(sql_statement)

with open("output.sql", "w", encoding='utf-8') as file:
    for v in values:
        file.write(v + '\n')
